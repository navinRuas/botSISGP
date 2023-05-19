import re
import pyodbc
import pandas as pd
import openpyxl
import os
import win32com.client as win32
from openpyxl.utils import get_column_letter

# Fazer a conexão com a base de dados
dadosconexao = (
    "Driver={SQL Server};"
    "Server=Pontalina.inep.gov.br;"
    "Database=PGD_SUSEP_PROD;"
    "Trusted_Connection=yes;"
)

# Verificar se a conexão deu certo
conexao = pyodbc.connect(dadosconexao)
print("conexão bem sucedida!")

df = pd.read_sql_query(f"SELECT NomeServidor, DtInicioPactoTrab, DtFimPactoTrab, DtInicioPactoTrabAtividade, DtFimPactoTrabAtividade, titulo, left(descricao, 500) as Descrição FROM [ProgramaGestao].[VW_PlanoTrabalhoAUDIN] where descricao like '%<demanda>%%</demanda>%<atividade>%%</atividade><produto>%%</produto><anoAcao>%%</anoAcao><idAcao>%%< idAcao><idSprint>%%</idSprint>%' and DtInicioPactoTrab BETWEEN DATEADD (DAY, 15, EOMONTH (GETDATE (), -2)) and GETDATE () and SituacaoPactoTrabalho != 'Executado' and SituacaoPactoTrabalho != 'Rejeitado'group by NomeServidor, DtInicioPactoTrab,DtFimPactoTrab, left(descricao, 500),DtInicioPactoTrabAtividade,DtFimPactoTrabAtividade, titulo order by NomeServidor, DtInicioPactoTrab", conexao)

# input = "<demanda>2</demanda><atividade>2</atividade><produto>2</produto><anoAcao>2023</anoAcao><idAcao>7</idAcao><idSprint></idSprint>"

def stripFunc(striptext, tagname):
    regex = r'<{0}>(.*?)<\/{0}>'.format(tagname)
    match = re.search(regex, striptext)

    if match:
        result = match.group(1)
        if (result == ""): result = "N/A"
        return result

def formatFile(file_path: str):
    # Read the file into a DataFrame
    df = pd.read_excel(file_path)

    # Update the column names
    df.columns = ["Nome Servidor", "Data Inicio Plano de Trabalho","Data Termino Plano de Trabalho","Data Inicio Atividade Plano de Trabalho","Data Termino Atividade Plano de Trabalho", "Atividade/Ação", "idEaud", "Demandas", "Atividades", "Produtos", "Ano", "Ação", "Sprint", "Descrição"]

    # Remove rows with duplicate values in column 'Descrição'
    df = df.drop_duplicates(subset=['Descrição'])

    # Write the formatted DataFrame back to the file
    df.to_excel(file_path, index=False)

def descTrans(input):
    demandas_dict = {
        "1": "Planejamento Anual",
        "2": "Avaliação",
        "3": "Consultoria",
        "4": "Apuração",
        "5": "Monitoramento",
        "6": "Demandas Externas",
        "7": "PGMQ",
        "8": "Demandas Administrativas",
        "9": "Demandas de TIC",
        "10": "Capacitação",
        "11": "Ausência",
        "12": "Participação em reuniões/GT",
        "13": "Outros"
    }

    ativades_dict = {
        "1": {
            "1": "Mapeamento do Universo de Auditoria",
            "2": "Elaboração/Atualização do PAINT",
            "3": "Elaboração/Atualização do RAINT"
        },
        "2": {
            "1": "Planejamento",
            "2": "Execução",
            "3": "Relatoria",
            "4": "Achados de Auditoria"
        },
        "3": {
            "1": "Planejamento",
            "2": "Execução",
            "3": "Relatoria",
            "4": "Achados de Auditoria"
        },
        "4": {
            "1": "Planejamento",
            "2": "Execução",
            "3": "Relatoria",
            "4": "Achados de Auditoria"
        },
        "5": {
            "1": "Monitoramento das Recomendações",
            "2": "Contabilização de benefícios"
        },
        "6": {
            "1": "Acompanhamento de Diligências TCU",
            "2": "Acompanhamento de Demandas CGU",
            "3": "Análise de admissibilidade de Denúncias",
            "4": "Suporte a Ação de Auditoria do TCU",
            "5": "Suporte a Ação de Auditoria da CGU"
        },
        "7": {
            "1": "Auto-Avaliação do IA-CM",
            "2": "Elaboração de Plano de Ação",
            "3": "Elaboração de Relatório de Avaliação IA-CM"
        },
        "8": {
        "1": "Gestão do SEI",
        "2": "Produção/Atualização de documentos"
    },
        "9": {
            "1": "Manipulação de Base de Dados",
            "2": "Desenvolvimento/Manutenção de Aplicativo",
            "3": "Desenvolvimento/Manutenção de Painel Gerencial",
            "4": "Gestão/Suporte e-Aud",
            "5": "Gestão do SharePoint",
            "6": "Outros"
        },
        "10": {
            "1": "Participação em cursos",
            "2": "Estudo individual"
        },
        "11": {
            "1": "Ausência"
        }
    }

    produtos_dict = {
        "1": {
            "1": {
                "1": "Universo de Auditoria"
            },
            "2": {
                "1": "PAINT Preliminar",
                "2": "PAINT Definitivo"
            },
            "3": {
                "1": "RAINT Preliminar",
                "2": "RAINT Definitivo"
            }
        },
        "2": {
            "1": {
                "1": "Análise Preliminar",
                "2": "Matriz de Riscos",
                "3": "Matriz de Planejamento"
            },
            "2": {
                "1": "Escopo da Auditoria",
                "2": "Papéis de Trabalho",
                "3": "Matriz de Achados"
            },
            "3": {
                "1": "Relatório Preliminar",
                "2": "Relatório Final"
            },
            "4": {
                "1": "Recomendações cadastradas"
            }
        },
        "3": {
            "1": {
                "1": "Análise Preliminar",
                "2": "Matriz de Riscos",
                "3": "Matriz de Planejamento"
            },
            "2": {
                "1": "Escopo da Auditoria",
                "2": "Papéis de Trabalho",
                "3": "Matriz de Achados"
            },
            "3": {
                "1": "Relatório Preliminar",
                "2": "Relatório Final"
            },
            "4": {
                "1": "Recomendações cadastradas"
            }
        },
        "4": {
            "1": {
                "1": "Análise Preliminar",
                "2": "Matriz de Riscos",
                "3": "Matriz de Planejamento"
            },
            "2": {
                "1": "Escopo da Auditoria",
                "2": "Papéis de Trabalho",
                "3": "Matriz de Achados"
            },
            "3": {
                "1": "Relatório Preliminar",
                "2": "Relatório Final"
            },
            "4": {
                "1": "Recomendações cadastradas"
            }
        },
        "5": {
            "1": {
                "1": "Recomendação monitorada"
            },
            "2": {
                "1": "Benefício contabilizado"
            }
        },
        "7": {
            "1": {
                "1": "Matriz de Avaliação",
                "2": "Avaliação IA-CM"
            },
            "2": {
                "1": "Plano de Ação"
            },
            "3": {
                "1": "Relatório de Avaliação IA-CM"
            }
        },
        "8": {
            "1": {
                "1": "Normativo",
                "2": "Parecer",
                "3": "Manual",
                "4": "POP"
            }
        },
        "9": {
            "1": {
                "1": "Consulta SQL"
            },
            "2": {
                "1": "Aplicativo"
            },
            "3": {
                "1": "Painel Gerencial"
            }
        },
        "11": {
            "1": {
                "1": "Atestado Médico",
                "2": "Férias"
            }
        }
    }

    acao_dict = {
        "01": "Ação 01/2022 - Elaboração de testes montagem de provas do Enem",
        "02": "Ação 02/2022 - Gerir Banco Nacional de Itens",
        "03": "Ação 03/2022 - Gestão da Integridade Pública",
        "03": "Ação 03/2022 - Processo de Montagem de testes do Enem",
        "07": "Ação 07/2022 - Processo de Concessão e Pagamentos da GECC",
        "08": "Ação 08/2022 - Gestão Orçamentária",
        "09": "Ação 09/2022 - Licitações e Contratos",
        "04": "Ação 04/2023 - Consultoria no Processo de Gestão de Riscos",
        "05": "Ação 05/2023 - Processos de Gestão da Contratação de Serviços Especializados de Aplicação do Enem/Desenvolver e Monitorar a Logistica dos Exames e valiação",
        "06": "Ação 06/2023 - Auditoria do Processo de Gestão do Banco de Dados de Especialistas",
        "07": "Ação 07/2023 - Auditoria do Portifólio de Projetos e Processos",
        "1": "Acompanhamento do PAINT",
        "2": "RAINT/2022",
        "3": "PAINT/2024",
        "4": "Acompanhamento/levantamento de auditorias CGU e TCU",
        "5": "Parecer sobre a prestação de contas anual do Inep",
        "6": "Supervisão",
        "7": "Monitoramento CGU/TCU",
        "8": "Monitoramento de recomendações",
        "9": "Gestão da Unidade",
        "10": "Gestão documental e controle de demandas externas."
    }

    # Fazer a conexão com a base de dados
    dadosconexao = (
        "Driver={SQL Server};"
        "Server=Pontalina.inep.gov.br;"
        "Database=PGD_SUSEP_PROD;"
        "Trusted_Connection=yes;"
    )

    # Verificar se a conexão deu certo
    conexao = pyodbc.connect(dadosconexao)
    print("conexão bem sucedida!")

    df = pd.read_sql_query(f"SELECT NomeServidor, DtInicioPactoTrab, DtFimPactoTrab, DtInicioPactoTrabAtividade, DtFimPactoTrabAtividade, titulo, left(descricao, 500) as Descrição FROM [ProgramaGestao].[VW_PlanoTrabalhoAUDIN] where descricao like '%<demanda>%%</demanda>%<atividade>%%</atividade><produto>%%</produto><anoAcao>%%</anoAcao><idAcao>%%< idAcao><idSprint>%%</idSprint>%' and DtInicioPactoTrab BETWEEN DATEADD (DAY, 15, EOMONTH (GETDATE (), -2)) and GETDATE () and SituacaoPactoTrabalho != 'Executado' and SituacaoPactoTrabalho != 'Rejeitado'group by NomeServidor, DtInicioPactoTrab,DtFimPactoTrab, left(descricao, 500),DtInicioPactoTrabAtividade,DtFimPactoTrabAtividade, titulo order by NomeServidor, DtInicioPactoTrab", conexao)

    # Creating calling Keys
    demandas_key = str(stripFunc(input, "demanda"))
    atividade_key = str(stripFunc(input, "atividade"))
    produto_key = str(stripFunc(input, "produto"))
    acao_key = str(stripFunc(input, "idAcao"))

    # Defining Excel file path
    file_path = 'Reverso.xlsx'  # Note: using .xls extension for compatibility with xlwt library

    # Verificando se o arquivo existe
    if os.path.exists(file_path):
        # Carregando o workbook existente
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.get_sheet_by_name('Sheet1') # Obtendo a referência da aba existente
        sheet = workbook.active
        #sheet.title = 'Principal'  # Renomeando a aba
    else:
        # Criando um novo workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active  # Selecionando a primeira aba
        sheet.title = 'Principal'  # Renomeando a aba

    # Find first empty cell in column
    row = 1 # Começando da primeira linha
    while sheet.cell(row=row, column=1).value is not  None: # Usando a função sheet.cell
        row += 1

    # Write data to cell
    df_filtrado = df.query ('Descrição == @input')
    
    # Verificando se a linha é menor ou igual a 1048576
    if row <= 1048576 and row > 0:
        # Escrevendo os dados na mesma planilha
        sheet.cell(row=row, column=1).value = df_filtrado.loc [:, 'NomeServidor'].to_string (index = False)
        sheet.cell(row=row, column=2).value = df_filtrado.loc [:, 'DtInicioPactoTrab'].to_string (index = False)
        sheet.cell(row=row, column=3).value = df_filtrado.loc [:, 'DtFimPactoTrab'].to_string (index = False)
        sheet.cell(row=row, column=4).value = df_filtrado.loc [:, 'DtInicioPactoTrabAtividade'].to_string (index = False)
        sheet.cell(row=row, column=5).value = df_filtrado.loc [:, 'DtFimPactoTrabAtividade'].to_string (index = False)
        sheet.cell(row=row, column=6).value = df_filtrado.loc [:, 'titulo'].to_string (index = False)
        sheet.cell(row=row, column=7).value = stripFunc(input, "idEaud")
        sheet.cell(row=row, column=8).value = demandas_dict.get(demandas_key, "N/A")
        sheet.cell(row=row, column=9).value = ativades_dict.get(demandas_key, {}).get(atividade_key, None)
        sheet.cell(row=row, column=10).value = produtos_dict.get(demandas_key, {}).get(atividade_key, {}).get(produto_key, None)

        sheet.cell(row=row, column=11).value = stripFunc(input, "anoAcao")
        sheet.cell(row=row, column=12).value = acao_dict.get(acao_key, "N/A")
        sheet.cell(row=row, column=13).value = stripFunc(input, "idSprint")
        sheet.cell(row=row, column=14).value = str(input)
        
        '''# Ajuste o tamanho da coluna para se ajustar ao conteúdo
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Obtenha a letra da coluna
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width'''

        # Save workbook
        workbook.save(file_path)
    
    formatFile("Reverso.xlsx")

for value in df['Descrição'].values:

    # Passar apenas o valor da coluna Descrição para a função descTrans
    descTrans(value)
    