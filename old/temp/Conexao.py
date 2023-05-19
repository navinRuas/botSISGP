import pandas as pd
import openpyxl
import pyodbc
import math

'''# Criando um dataframe de exemplo
df = pd.DataFrame({"titulo": ["T1", "T2", "T3"], "descricao": ["<demanda>2</demanda><atividade>1</atividade><produto>3</produto>", "<demanda>1</demanda><atividade>2</atividade><produto>1</produto>", "<demanda>3</demanda><atividade>3</atividade><produto>2</produto>"]})'''

# Fazer a conexão com a base de dados
dadosconexao = (
    "Driver={SQL Server};"
    "Server=Pontalina.inep.gov.br;"
    "Database=PGD_SUSEP_PROD;"
    "Trusted_Connection=yes;"
)

# Verificar se a conexão deu certo
conexao = pyodbc.connect(dadosconexao)
print("conexão bem sucedida!")

#área onde acontecerar a pesquisa
df = pd.read_sql_query(f"SELECT NomeServidor, DtInicioPactoTrab, titulo as Título, left(descricao, 200) as Descrição FROM [ProgramaGestao].[VW_PlanoTrabalhoAUDIN] where descricao like '%<demanda>%%</demanda>%<atividade>%%</atividade><produto>%%</produto><anoAcao>%%</anoAcao><idAcao>%%</idAcao><idSprint>%%</idSprint>%' and DtInicioPactoTrab BETWEEN DATEADD (DAY, 1, EOMONTH (GETDATE (), -1)) and GETDATE () and SituacaoPactoTrabalho != 'Executado' and SituacaoPactoTrabalho != 'Rejeitado' group by NomeServidor, DtInicioPactoTrab, left(descricao, 200), titulo order by NomeServidor, DtInicioPactoTrab", conexao)
#print(df)

# Lendo a planilha excel como um dataframe
planilha = pd.read_excel("C://Users\jamil.monteiro\OneDrive - INEP\Documents\Projeto\Site\De-Para codificada.xlsx", header=1)
#print(planilha['CodDemanda'].values)
#print(planilha['CodAtividade'].values)
#print(planilha['CodProduto'].values)
#print(planilha)

# Definindo uma função que verifica se a descrição corresponde às colunas da planilha
def verificar_descricao(descricao):
  # Extraindo a parte da descrição que interessa
  descricao = descricao.split("</produto>")[0] + "</produto>"
  # Extraindo os valores de demanda, atividade e produto da descricao
  demanda = descricao.split("<demanda>")[-1].split("</demanda>")[0]
  atividade = descricao.split("<atividade>")[-1].split("</atividade>")[0]
  produto = descricao.split("<produto>")[-1].split("</produto>")[0]
  
  # Lendo o arquivo excel em um dataframe
  dfs = pd.read_excel("C://Users\jamil.monteiro\OneDrive - INEP\Documents\Projeto\Site\De-Para codificada.xlsx", sheet_name="de-para",header=1)
  # Criando uma coluna com a concatenação de demanda, atividade e produto
  dfs["chave"] = dfs["CodDemanda"].astype(str) + "&" + dfs["CodAtividade"].astype(str) + "&" + dfs["CodProduto"].astype(str)
  # Criando um dicionário que mapeia a chave para os valores das colunas 1,4 e 7
  dic = dfs.set_index("index")[["CodDemanda", "CodAtividade", "CodProduto"]].to_dict(orient="index")
  # Criando um dicionário para pegar os valores das colunas 10 e 11
  ret = dfs.set_index("index")[["Atividade2", "nº da atividade"]].to_dict(orient="index")
  # Criando a chave a partir dos valores de demanda, atividade e produto da descrição
  chave = '&'.join([str(demanda), str(atividade), str(produto)])
  # Criando uma lista vazia para armazenar as novas chaves de dic
  novas_chaves = []
  new_key = []

  for k in ret:
    atividade2 = ret[k]["Atividade2"]
    numero = ret[k]["nº da atividade"]
    
    if math.isnan(numero):
        numero = 0
    else:
        numero = int(numero)
    
    key = '-'.join([str(atividade2), str(numero)])
    new_key.append(key)

  # Percorrendo as chaves de dic
  for k in dic:
    # Obtendo os valores de demanda, atividade e produto de cada chave
    demanda = dic[k]['CodDemanda']
    atividade = dic[k]['CodAtividade']
    produto = dic[k]['CodProduto']
    
    # Verificando se os valores são NaN
    if math.isnan(demanda):
        demanda = 0
    else:
        demanda = int(demanda)
    
    if math.isnan(atividade):
        atividade = 0
    else:
        atividade = int(atividade)
    
    if math.isnan(produto):
        produto = 0
    else:
        produto = int(produto)
    
    # Concatenando os valores com o caractere '&' e adicionando à lista de novas chaves
    nova_chave = '&'.join([str(demanda), str(atividade), str(produto)])
    novas_chaves.append(nova_chave)
    
  df6 = df['Título'].str.slice(0,6)
  # Verificando se a chave existe no dicionário
  if chave in novas_chaves:
    valor = chave
    if key in df6: 
      print("Key: ",key)
      #print("New_key: ",new_key)
      #print("Chave: ", chave)
      #print("Dic: ", dic)
      #print("Ret:", ret)
      #print(df["Título"] == key)
      # Retornando o valor correspondente
      valor = chave
      #print(valor)
  else:
      # Retornando um valor padrão
      valor = "Não encontrado"
      #print(valor)
  return valor
  # Extraindo a parte da descrição que interessa
  descricao = descricao.split("</produto>")[0] + "</produto>"
  # Extraindo os valores de demanda, atividade e produto da descricao
  demanda = int(descricao.split("<demanda>")[-1].split("</demanda>")[0])
  atividade = int(descricao.split("<atividade>")[-1].split("</atividade>")[0])
  produto = int(descricao.split("<produto>")[-1].split("</produto>")[0])
  print("Esse é o valor entre <demanda></demanda>",demanda)
  print("Esse é o valor entre <atividade></atividade>",atividade)
  print("Esse é o valor entre <produto></produto>",produto)

  # Lendo o arquivo excel
  wb = openpyxl.load_workbook("C://Users\jamil.monteiro\OneDrive - INEP\Documents\Projeto\Site\De-Para codificada.xlsx", data_only=True)
  # Acessando a planilha desejada
  ws = wb["de-para"]
  # Aplicando a fórmula na célula D67
  ws.cell(row=67, column=4).value = '=ÍNDICE(A:K;CORRESP({}&{}&{};A:A&D:D&G:G;0);"{8;10;11})'.format(demanda, atividade, produto)
  # Pegando o valor da célula D67
  valor = str(ws.cell(row=67, column=4).value) + '-' + str(ws.cell(row=68, column=4).value) + '-' + str(ws.cell(row=69, column=4).value)
  # Imprimindo o valor
  print(valor)
  return(valor)

resultado = df["Descrição"].apply(verificar_descricao)
'''print(resultado)
if any(df['Título'].values) == resultado:
  print()'''

# Imprimindo o resultado
#print(resultado)
#Código para o excel
#=ÍNDICE (A:D; CORRESP ("<demanda>2</demanda><atividade>1</atividade><produto>3</produto>"; A:A & B:B & C:C; 0); 4)
